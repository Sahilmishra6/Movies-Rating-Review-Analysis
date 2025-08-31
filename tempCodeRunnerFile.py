import os
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# 1) Load movie data
def load_movie_data(file_path: str):
    if 'Movies rating & review data' in pd.ExcelFile(file_path).sheet_names:
        df_movies = pd.read_excel(file_path, sheet_name='Movies rating & review data')
    else:
        df_movies = pd.DataFrame()
    return df_movies

# 2) Clean data
def clean_movie_data(df_movies):
    df_movies = df_movies.fillna({
        'Movie_Name': 'Unknown',
        'Release_Year': 0,
        'Genre': 'Unknown',
        'Ratings': 0,
        'Reviewer': 'Unknown',
        'Reviews': 'Unknown',
        'Directors': 'Unknown',
        'Writers': 'Unknown'
    })
    df_movies = df_movies.drop_duplicates()
    return df_movies

# 3) Analyze movie data
def analyze_movies(df_movies):
    genre_dist = df_movies['Genre'].value_counts()
    ratings_stats = pd.Series({
        'Average_Rating': df_movies['Ratings'].mean(),
        'Min_Rating': df_movies['Ratings'].min(),
        'Max_Rating': df_movies['Ratings'].max()
    })
    return genre_dist, ratings_stats

# 4) Optional: analyze reviews/preferences
def analyze_movies_years(df_movies):
    # Count movies per year
    movie_counts_series = df_movies['Release_Year'].value_counts().sort_index()

    # Find year with max count
    max_year = movie_counts_series.idxmax()
    max_count = movie_counts_series.max()

    # Return formatted summary (year + count)
    movie_summary = f"{max_year} ({max_count} movies)"

    return movie_counts_series, movie_summary

# 5) Export report
def export_report(output_file, genre_dist, ratings_stats, movie_counts):
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        genre_dist.rename_axis('Genre').reset_index(name='Count').to_excel(writer, sheet_name='Genre_Wise_movies', index=False)
        ratings_stats.rename('Value').reset_index().rename(columns={'index':'Metric'}).to_excel(writer, sheet_name='Ratings_Stats', index=False)
        movie_counts.rename_axis('Release Year').reset_index(name='Movies').to_excel(writer, sheet_name='Release_year_Movies', index=False)

# 6) Generate charts
def generate_charts(df_movies, charts_dir='charts'):
    os.makedirs(charts_dir, exist_ok=True)
    paths = {}
    sns.set_theme(style='darkgrid')
    
    # Genre pie
    plt.figure(figsize=(6,6))
    df_movies['Genre'].value_counts().plot.pie(autopct='%1.1f%%')
    plt.title('Genre Distribution')
    genre_path = os.path.join(charts_dir, 'genre_pie.png')
    plt.savefig(genre_path)
    plt.close()
    paths['genre'] = genre_path

    # Rating histogram
    plt.figure(figsize=(8,5))
    sns.histplot(df_movies['Ratings'], bins=10, kde=True)
    plt.title('Ratings Distribution')
    rating_path = os.path.join(charts_dir, 'ratings_hist.png')
    plt.savefig(rating_path)
    plt.close()
    paths['ratings'] = rating_path

    # Movies per year bar chart
    df_movies['Release_Year'] = df_movies['Release_Year'].fillna(0).astype(int)
    df_movies_filtered = df_movies[df_movies['Release_Year'] != 0]  # remove placeholder 0

    year_counts = df_movies_filtered.groupby('Release_Year')['Movie_Name'].count().sort_index()

    plt.figure(figsize=(10,6))
    sns.barplot(x=year_counts.index, y=year_counts.values, palette="viridis")
    plt.xlabel('Release Year')
    plt.ylabel('Number of Movies')
    plt.title('Movies Released Per Year')
    plt.xticks(rotation=45, ha='right')

    release_year_path = os.path.join(charts_dir, 'Release_year_movies_bar.png')
    plt.savefig(release_year_path)
    plt.close()
    paths['Release_Year'] = release_year_path

    return paths

# 7) Embed charts into Excel
def embed_charts(excel_file, chart_paths):
    wb = load_workbook(excel_file)
    if 'Charts' not in wb.sheetnames:
        wb.create_sheet('Charts')
    ws = wb['Charts']
    row = 1
    for label, path in chart_paths.items():
        img = XLImage(path)
        ws.add_image(img, f'A{row}')
        row += 20
    wb.save(excel_file)

# ---------------- Orchestration ---------------- #
def main():
    source_file = 'movie_data.xlsx'
    output_excel = 'movies_analysis_report_streamlit.xlsx'
    charts_dir = 'charts'

    # Load movie data
    df_movies = load_movie_data(source_file)

    # Clean data
    df_movies = clean_movie_data(df_movies)

    # Analyze movies
    genre_dist, ratings_stats = analyze_movies(df_movies)
    movie_counts, movie_summary = analyze_movies_years(df_movies)

    # Export report
    export_report(output_excel, genre_dist, ratings_stats, movie_counts)

    # Generate charts and embed into Excel
    chart_paths = generate_charts(df_movies, charts_dir)
    embed_charts(output_excel, chart_paths)

    # Console summary
    print("✅ Analysis Completed!\n")
    print("Most movies Released in year:", movie_summary)
    print("\nNumber of movies per genre:")
    print(genre_dist)
    print("\nRating stats:")
    print(ratings_stats)
    print("\nCharts saved:")
    for k, v in chart_paths.items():
        print(f" - {k}: {v}")
    print(f"\nReport written to: {output_excel}")


if __name__ == '__main__':
    main()
